"""Convierte la hoja 'Journals' del Excel JCR a CSV listo para importar en Supabase."""
import csv
import sys
from pathlib import Path
import openpyxl

EXCEL = Path(__file__).parent / "jcr_update.xlsx"
OUTPUT = Path(__file__).parent / "journals_clean.csv"

COLUMN_MAP = {
    "ISSN": "issn",
    "EISSN": "eissn",
    "TITLE": "title",
    "TITLE20": "title_abbrev",
    "ISO_ABBREV": "iso_abbrev",
    "YEAR": "year",
    "IMFACT_FACTOR": "impact_factor",
    "IMPACT_FACTOR_5YR": "impact_factor_5yr",
    "EIGENFACTOR": "eigenfactor",
    "ARTL_INFLUENCE": "article_influence",
    "IMMEDIACY_INDEX": "immediacy_index",
    "NORM_EIGENFACTOR": "norm_eigenfactor",
    "QUARTILE_RANK": "quartile_rank",
    "JIF_PERCENTILE": "jif_percentile",
    "CATEGORY_RANKING": "category_ranking",
    "CATEGORIES_CODE": "categories_code",
    "CATEGORIES_DESCRIPTION": "categories_description",
    "EDITION": "edition",
    "PUBLISHER_NAME": "publisher_name",
    "COUNTRY": "country",
    "ADDRESS": "address",
    "TOT_CITES": "total_cites",
    "CITED_HALF_LIFE": "cited_half_life",
}

NULL_STRINGS = {"NULL", "null", "N/A", "n/a", "NA", "na", ""}


def clean(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    return "" if s in NULL_STRINGS else s


def clean_issn(value) -> str:
    s = clean(value)
    if len(s) == 8 and s.isdigit():
        return f"{s[:4]}-{s[4:]}"
    return s


print(f"Abriendo {EXCEL}...")
wb = openpyxl.load_workbook(EXCEL, read_only=True)
ws = wb["Journals"] if "Journals" in wb.sheetnames else wb.active
print(f"Hoja: '{ws.title}' (~{ws.max_row} filas)")

rows_iter = ws.iter_rows(values_only=True)
headers = [str(c).strip() if c else "" for c in next(rows_iter)]
col_idx = {col: i for i, col in enumerate(headers) if col in COLUMN_MAP}
print(f"Columnas mapeadas: {len(col_idx)}/{len(COLUMN_MAP)}")

db_cols = list(COLUMN_MAP.values())

written = 0
skipped = 0

with open(OUTPUT, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(db_cols)

    for row in rows_iter:
        title_idx = col_idx.get("TITLE")
        title = clean(row[title_idx]) if title_idx is not None else ""
        if not title:
            skipped += 1
            continue

        out = []
        for excel_col, db_col in COLUMN_MAP.items():
            if excel_col not in col_idx:
                out.append("")
                continue
            val = row[col_idx[excel_col]]
            if db_col in ("issn", "eissn"):
                out.append(clean_issn(val))
            else:
                out.append(clean(val))

        writer.writerow(out)
        written += 1

wb.close()
print(f"\nListo: {written:,} filas escritas → {OUTPUT}")
print(f"Omitidas (sin título): {skipped}")
